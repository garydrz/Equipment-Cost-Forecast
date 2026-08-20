"""CSV / Excel import + export utilities for the EPC estimator.

The helpers here are IO-only. They receive/return primitive Python dicts;
the FastAPI route layer is responsible for reading the request body,
validating documents and calling into `db.py` for persistence.
"""
import csv
import io
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional


HISTORICAL_COLUMNS = [
    "id", "category", "subtype", "size", "size_unit", "weight_kg", "material",
    "design_pressure_bar", "design_temperature_c", "power_kw",
    "flow_rate_m3_h", "head_m", "pump_efficiency", "fluid_density_kg_m3",
    "thermal_duty_kw", "fuel_flow_kg_h",
    "year", "cost_original", "currency", "vendor_country", "install_country", "notes",
]

ROW_COLUMNS = [
    "id", "project_id", "tag", "category", "subtype", "size", "size_unit",
    "weight_kg", "material", "design_pressure_bar", "design_temperature_c",
    "power_kw", "flow_rate_m3_h", "head_m", "pump_efficiency",
    "fluid_density_kg_m3", "thermal_duty_kw", "fuel_flow_kg_h",
    "quantity", "unit_expected_cost", "unit_low", "unit_high",
    "total_expected_cost", "total_sigma",
]

PROJECT_COLUMNS = [
    "id", "name", "description", "output_currency", "target_year", "aace_class", "created_at",
]


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------
def parse_csv(text: str) -> List[Dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for r in reader:
        # Strip empty strings -> None, and cast numerics best-effort
        cleaned: Dict[str, Any] = {}
        for k, v in r.items():
            if k is None:
                continue
            key = k.strip()
            if v is None or v == "":
                cleaned[key] = None
            else:
                cleaned[key] = _coerce(v)
        rows.append(cleaned)
    return rows


def to_csv(records: Iterable[Dict[str, Any]], columns: List[str]) -> str:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    w.writeheader()
    for r in records:
        row = {c: r.get(c, "") for c in columns}
        # Flatten dicts/lists so Excel opens them cleanly
        for k, v in row.items():
            if isinstance(v, (dict, list)):
                row[k] = ""
            elif v is None:
                row[k] = ""
        w.writerow(row)
    return buf.getvalue()


def _coerce(v: str) -> Any:
    v = v.strip()
    if v.lower() in ("true", "false"):
        return v.lower() == "true"
    # int?
    try:
        if v.isdigit() or (v.startswith("-") and v[1:].isdigit()):
            return int(v)
    except AttributeError:
        pass
    # float?
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


# ---------------------------------------------------------------------------
# Historical import
# ---------------------------------------------------------------------------
def normalize_historical(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Return a dict suitable for HistoricalEquipmentCreate.model_validate."""
    def num(k: str, cast=float) -> Optional[Any]:
        v = raw.get(k)
        if v in (None, "", "None"):
            return None
        try:
            return cast(v)
        except (TypeError, ValueError):
            return None

    out = {
        "category": (raw.get("category") or "").strip(),
        "subtype": (raw.get("subtype") or "").strip(),
        "material": (raw.get("material") or "").strip(),
        "currency": (raw.get("currency") or "EUR").strip().upper(),
        "size": num("size", float) or 0,
        "size_unit": raw.get("size_unit") or None,
        "weight_kg": num("weight_kg", float),
        "design_pressure_bar": num("design_pressure_bar", float),
        "design_temperature_c": num("design_temperature_c", float),
        "power_kw": num("power_kw", float),
        "flow_rate_m3_h": num("flow_rate_m3_h", float),
        "head_m": num("head_m", float),
        "pump_efficiency": num("pump_efficiency", float),
        "fluid_density_kg_m3": num("fluid_density_kg_m3", float),
        "thermal_duty_kw": num("thermal_duty_kw", float),
        "fuel_flow_kg_h": num("fuel_flow_kg_h", float),
        "year": num("year", int) or datetime.now(timezone.utc).year,
        "cost_original": num("cost_original", float) or 0.0,
        "vendor_country": raw.get("vendor_country") or None,
        "install_country": raw.get("install_country") or None,
        "notes": raw.get("notes") or None,
    }
    return out


# ---------------------------------------------------------------------------
# Excel export (project workbook)
# ---------------------------------------------------------------------------
def project_to_excel_bytes(project: Dict[str, Any], rows: List[Dict[str, Any]],
                           totals: Dict[str, Any]) -> bytes:
    """Emit a multi-sheet XLSX with project overview, rows and totals.

    Requires openpyxl. Kept in this module (not the routes) so it's simple
    to unit test with a plain dict input.
    """
    from openpyxl import Workbook

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Project"
    ws1.append(["field", "value"])
    for k in ("id", "name", "description", "output_currency", "target_year", "aace_class", "created_at"):
        v = project.get(k)
        if isinstance(v, datetime):
            v = v.isoformat(timespec="seconds")
        ws1.append([k, v])

    ws2 = wb.create_sheet("Equipment Rows")
    ws2.append(ROW_COLUMNS)
    for r in rows:
        ws2.append([_scalar(r.get(c)) for c in ROW_COLUMNS])

    ws3 = wb.create_sheet("Totals")
    ws3.append(["field", "value"])
    for k, v in (totals or {}).items():
        ws3.append([k, _scalar(v)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def historical_to_excel_bytes(records: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Equipment"
    ws.append(HISTORICAL_COLUMNS)
    for r in records:
        ws.append([_scalar(r.get(c)) for c in HISTORICAL_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _scalar(v: Any) -> Any:
    if isinstance(v, (dict, list)):
        return ""
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    return v
